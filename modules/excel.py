from openpyxl import Workbook
from openpyxl import load_workbook
import os


class Excel:
    def __init__(self, file):
        self.file = file
        self.is_exists = False

        if os.path.exists(file):
            self.wb = load_workbook(file)
            self.is_exists = True
        else:
            self.wb = Workbook()

        self.ws = self.wb.active

    def write_header(self, headers):
        if not self.is_exists:
            self.ws.append(headers)

    def write_rows(self, rows):
        for row in rows:
            self.ws.append(row)

    def save(self):
        self.wb.save(self.file)
