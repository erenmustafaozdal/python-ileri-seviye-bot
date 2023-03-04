from openpyxl import Workbook
from openpyxl import load_workbook
import os

# Excel dosyasını oluşturalım
excel_file = "linkedin_posts.xlsx"
if os.path.exists(excel_file):  # eğer dosya varsa
    wb = load_workbook(excel_file)
    ws = wb.active
else:  # eğer dosya yoksa
    wb = Workbook()
    ws = wb.active
    ws.append([
        "Owner Name",
        "Owner URL",
        "Date",
        "Text",
        "Shared Text",
    ])


wb.save(excel_file)
